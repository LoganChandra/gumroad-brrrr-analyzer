#!/usr/bin/env python3
"""Build the BRRRR Deal Analyzer .xlsx spreadsheet with all tabs."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from copy import copy

# ── Style Constants ──────────────────────────────────────────────────────────
DARK_NAVY = "1B2A4A"
ACCENT_GREEN = "27AE60"
ACCENT_RED = "E74C3C"
ACCENT_YELLOW = "F39C12"
LIGHT_GRAY = "F5F6FA"
WHITE = "FFFFFF"
BLACK = "333333"
BORDER_GRAY = "D0D5DD"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, color=DARK_NAVY, size=11)
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
green_fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")
red_fill = PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid")
yellow_fill = PatternFill(start_color=ACCENT_YELLOW, end_color=ACCENT_YELLOW, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)
bottom_border = Border(bottom=Side(style="medium", color=DARK_NAVY))

label_font = Font(name="Calibri", bold=True, color=BLACK, size=11)
value_font = Font(name="Calibri", color=BLACK, size=11)
value_font_dollar = Font(name="Calibri", color=BLACK, size=11)
pct_font = Font(name="Calibri", color=BLACK, size=11)
title_font = Font(name="Calibri", bold=True, color=DARK_NAVY, size=16)
section_font = Font(name="Calibri", bold=True, color=DARK_NAVY, size=13)
note_font = Font(name="Calibri", italic=True, color="888888", size=10)
score_display_font = Font(name="Calibri", bold=True, color=WHITE, size=18)


def apply_cell(ws, row, col, value, font=None, fill=None, alignment=None,
               number_format=None, border=None, width=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border
    else:
        cell.border = thin_border
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def set_header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        apply_cell(ws, row, i, h, font=header_font, fill=header_fill,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                   width=widths[i-1] if widths else 18)


def set_subheader_row(ws, row, items):
    for i, item in enumerate(items, 1):
        apply_cell(ws, row, i, item, font=subheader_font, fill=subheader_fill)


def set_label(ws, row, col, label, width=None, note=None):
    apply_cell(ws, row, col, label, font=label_font, fill=light_fill,
               alignment=Alignment(horizontal="left", vertical="center"),
               width=width)
    if note:
        apply_cell(ws, row, col + 2, note, font=note_font, fill=white_fill,
                   alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    return ws.cell(row=row, column=col)


def set_value(ws, row, col, value, fmt=None, bold=False):
    f = Font(name="Calibri", bold=bold, color=BLACK, size=12)
    cell = apply_cell(ws, row, col, value, font=f,
                      alignment=Alignment(horizontal="center", vertical="center"),
                      number_format=fmt)
    return cell


def set_big_label(ws, row, col, text, width=None):
    apply_cell(ws, row, col, text, font=Font(name="Calibri", bold=True, color=DARK_NAVY, size=14),
               fill=white_fill,
               alignment=Alignment(horizontal="left", vertical="center"),
               width=width)


def merge_title(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = title_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 35


def merge_section(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = section_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = subheader_fill


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Deal Input
# ═══════════════════════════════════════════════════════════════════════════════
def build_input_sheet(wb):
    ws = wb.active
    ws.title = "Deal Input"
    ws.sheet_properties.tabColor = DARK_NAVY

    col_widths = [40, 20, 30, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    merge_title(ws, 1, 1, 4, "🏠 BRRRR & Rental Deal Analyzer — Input Sheet")
    ws.row_dimensions[1].height = 40

    ln = 3
    merge_section(ws, ln, 1, 4, "PROPERTY & PURCHASE")
    ln = 5
    inputs = [
        ("Property Address", "123 Main St, Anytown, USA", "$0.00", None),
        ("Purchase Price", None, "$250,000", "$0"),
        ("After-Repair Value (ARV)", None, "$350,000", "$0"),
        ("Rehab / Renovation Budget", None, "$40,000", "$0"),
        ("Holding Period (months)", None, "6", "0"),
        ("Closing Costs (% of Purchase)", None, "6%", "0%"),
    ]
    for label, note, default, fmt in inputs:
        set_label(ws, ln, 1, label)
        cell = set_value(ws, ln, 2, default)
        if note:
            set_label(ws, ln, 3, note)
        ln += 1

    ln += 1
    merge_section(ws, ln, 1, 4, "FINANCING")
    ln += 2
    fin_inputs = [
        ("Down Payment (%)", "20%", "0%"),
        ("Interest Rate (Annual)", "7.5%", "0%"),
        ("Loan Term (Years)", "30", "0"),
        ("Refinance LTV (%)", "70%", "0%"),
        ("Refinance Interest Rate", "7.0%", "0%"),
        ("Refinance Closing Costs", "$5,000", "$0"),
    ]
    for label, default, fmt in fin_inputs:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, default)
        ln += 1

    ln += 1
    merge_section(ws, ln, 1, 4, "RENTAL INCOME & EXPENSES")
    ln += 2
    rent_inputs = [
        ("Monthly Rent (After Rehab)", "$2,800", "$0"),
        ("Property Management (5% of Rent)", "5%", "0%"),
        ("Vacancy Rate", "5%", "0%"),
        ("Property Tax (Annual)", "$3,600", "$0"),
        ("Insurance (Annual)", "$1,200", "$0"),
        ("HOA (Monthly)", "$0", "$0"),
        ("Maintenance Reserve (% of Rent)", "10%", "0%"),
        ("Capex Reserve (% of Rent)", "5%", "0%"),
        ("Other Monthly Expenses", "$100", "$0"),
    ]
    for label, default, fmt in rent_inputs:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, default)
        ln += 1

    ln += 1
    merge_section(ws, ln, 1, 4, "FLIP SCENARIO (Separate Tab)")
    ln += 2
    flip_inputs = [
        ("Flip Sale Price", "$365,000", "$0"),
        ("Seller Closing Costs", "3%", "0%"),
        ("Real Estate Commission", "5%", "0%"),
    ]
    for label, default, fmt in flip_inputs:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, default)
        ln += 1

    # Note at bottom
    ln += 2
    ws.merge_cells(start_row=ln, start_column=1, end_row=ln, end_column=4)
    cell = ws.cell(row=ln, column=1,
                   value="💡 Yellow cells are editable inputs. Change values to run your deal.")
    cell.font = Font(name="Calibri", italic=True, color="888888", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: BRRRR Analyzer
# ═══════════════════════════════════════════════════════════════════════════════
def build_brrrr_analyzer(wb):
    ws = wb.create_sheet("BRRRR Analyzer")
    ws.sheet_properties.tabColor = ACCENT_GREEN

    col_widths = [40, 22, 22, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    merge_title(ws, 1, 1, 4, "💰 BRRRR Analysis — Buy → Rehab → Rent → Refinance → Repeat")
    ws.row_dimensions[1].height = 40

    ln = 3
    merge_section(ws, ln, 1, 4, "TOTAL CASH NEEDED (Before Refinance)")
    ln += 2

    items = [
        ("Down Payment", '=ROUND("Deal Input"!C7*"Deal Input"!C13,2)', "down_payment"),
        ("Closing Costs (Purchase)", '=ROUND("Deal Input"!C7*"Deal Input"!C11,2)', "closing"),
        ("Rehab Budget", '="Deal Input"!C9', "rehab"),
        ("Holding Costs (6 months × PITI + utils)", '=ROUND(("Deal Input"!C7-"Deal Input"!C13*"Deal Input"!C7)/"Deal Input"!C15/12*"Deal Input"!C10,2)', "holding"),
        ("Upfront Total Investment", '=SUM(C5:C8)', "total"),
    ]
    for label, formula, ref in items:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, formula)
        if ref == "total":
            ws.cell(row=ln, column=2).font = Font(name="Calibri", bold=True, color=DARK_NAVY, size=13)
        ln += 1

    ln += 2
    merge_section(ws, ln, 1, 4, "REFINANCE PROCEEDS — KEY FORMULA")
    ln += 2
    set_label(ws, ln, 1, "ARV × Refi LTV − Purchase × 1.06 (Closing) − Rehab − Holding × 6 Months")
    set_value(ws, ln, 2, "= CASH YOU PULL OUT")
    ws.cell(row=ln, column=2).font = Font(name="Calibri", bold=True, color=ACCENT_GREEN, size=12)
    ln += 2

    refi_items = [
        ("After-Repair Value (ARV)", '="Deal Input"!C8', "arv"),
        ("Refinance LTV %", '="Deal Input"!C16', "ltv"),
        ("Max Refi Loan Amount", '=ROUND(B14*B15,2)', "max_loan"),
        ("Purchase Price (incl. 6% closing)", '=ROUND("Deal Input"!C7*1.06,2)', "purchase_adj"),
        ("Rehab Costs", '="Deal Input"!C9', "rehab2"),
        ("Holding Costs (6 months)", '=C7', "holding2"),
        ("Cash Pulled Out (Refi Proceeds − Costs)", '=B16-B17-B18-B19', "cash_out"),
        ("Cash Left In (If Negative: You're bringing cash)", '=C9-B20', "cash_left"),
    ]
    for label, formula, ref in refi_items:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, formula)
        if ref in ("cash_out", "cash_left"):
            ws.cell(row=ln, column=2).font = Font(name="Calibri", bold=True, color=ACCENT_GREEN, size=13)
            # Conditional formatting note
            ws.cell(row=ln, column=3).value = "✅ Green = positive cash-out"
            ws.cell(row=ln, column=3).font = note_font
        ln += 1

    ln += 2
    merge_section(ws, ln, 1, 4, "CASH-ON-CASH RETURN (Pre-Refi)")
    ln += 2
    coc_items = [
        ("Annual Net Operating Income", '=ROUND(("Deal Input"!C20*12)*(1-"Deal Input"!C21-"Deal Input"!C22)-"Deal Input"!C23-"Deal Input"!C24-"Deal Input"!C26*12,2)', "noi"),
        ("Annual Debt Service (P&I)", '=ROUND(PMT("Deal Input"!C14/12,"Deal Input"!C15*12,-("Deal Input"!C7-"Deal Input"!C13*"Deal Input"!C7)),2)*12', "debt"),
        ("Annual Cash Flow", '=ROUND(B27-B28,2)', "cf"),
        ("Cash on Cash Return", '=IF(C9>0,B29/C9,"N/A")', "coc"),
    ]
    for label, formula, ref in coc_items:
        set_label(ws, ln, 1, label)
        if ref == "coc":
            set_value(ws, ln, 2, formula, fmt="0.00%")
        else:
            set_value(ws, ln, 2, formula, fmt="$#,##0.00")
        ln += 1

    ln += 2
    merge_section(ws, ln, 1, 4, "BRRRR DEAL QUALITY SCORE")
    ln += 2
    score_items = [
        ("Cash Pulled Out", '=IF(B20>0,"Positive ✅","Negative ❌")', ""),
        ("CoC Return Score", '=IF(B30>=0.1,"Green ✅",IF(B30>=0.05,"Yellow","Red ❌"))', ""),
        ("Overall Score", '=IF(AND(B20>0,B30>=0.1),"GREEN LIGHT 🟢 — BRRRR this deal!",IF(OR(B20>0,B30>=0.05),"YELLOW ⚠️ — Proceed carefully","RED 🔴 — Pass on this deal"))', ""),
    ]
    for label, formula, _ in score_items:
        set_label(ws, ln, 1, label)
        cell = set_value(ws, ln, 2, formula)
        ln += 1

    ln += 2
    cell = ws.cell(row=ln, column=1,
                   value="Note: The KEY FORMULA is ARV × Refi LTV − Purchase × 1.06 (Closing) − Rehab − Holding × 6 months = Cash you actually pull out.")
    cell.font = Font(name="Calibri", italic=True, color="888888", size=10)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Flip Analyzer
# ═══════════════════════════════════════════════════════════════════════════════
def build_flip_analyzer(wb):
    ws = wb.create_sheet("Flip Analyzer")
    ws.sheet_properties.tabColor = ACCENT_YELLOW

    col_widths = [40, 22, 22, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    merge_title(ws, 1, 1, 4, "🔨 Flip Analysis — Purchase → Rehab → Sell")
    ws.row_dimensions[1].height = 40

    ln = 3
    merge_section(ws, ln, 1, 4, "ACQUISITION COSTS")
    ln += 2
    acq_items = [
        ("Purchase Price", '="Deal Input"!C7', ""),
        ("Closing Costs (Purchase)", '=ROUND("Deal Input"!C7*"Deal Input"!C11,2)', ""),
        ("Rehab Budget", '="Deal Input"!C9', ""),
        ("Total Acquisition", '=SUM(C4:C6)', ""),
    ]
    for label, formula, _ in acq_items:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, formula)
        ln += 1

    ln += 2
    merge_section(ws, ln, 1, 4, "SELLING COSTS & REVENUE")
    ln += 2
    sell_items = [
        ("Flip Sale Price", '="Deal Input"!C32', ""),
        ("Seller Closing Costs (3%)", '=ROUND("Deal Input"!C32*"Deal Input"!C33,2)', ""),
        ("Real Estate Commission (5%)", '=ROUND("Deal Input"!C32*"Deal Input"!C34,2)', ""),
        ("Total Selling Costs", '=SUM(C12:C13)', ""),
        ("Net Sale Proceeds", '=C11-C14', ""),
    ]
    for label, formula, _ in sell_items:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, formula)
        ln += 1

    ln += 2
    merge_section(ws, ln, 1, 4, "PROFIT & RETURN")
    ln += 2
    profit_items = [
        ("Holding Costs (PITI)", '="BRRRR Analyzer"!C7', ""),
        ("Total All-In Cost", '=C7+C18', ""),
        ("Gross Profit", '=C15-C19', ""),
        ("Return on Investment (ROI)", '=IF(C19>0,C20/C19,"N/A")', "roi"),
        ("Annualized ROI (6-month hold)", '=IF(C21>0,C21*2,"N/A")', "aroi"),
    ]
    for label, formula, ref in profit_items:
        set_label(ws, ln, 1, label)
        if ref in ("roi", "aroi"):
            set_value(ws, ln, 2, formula, fmt="0.00%")
        else:
            set_value(ws, ln, 2, formula, fmt="$#,##0.00")
        ln += 1

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Buy and Hold (30yr Projection)
# ═══════════════════════════════════════════════════════════════════════════════
def build_buy_and_hold(wb):
    ws = wb.create_sheet("Buy & Hold 30yr")
    ws.sheet_properties.tabColor = "2E86C1"

    col_widths = [18, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    merge_title(ws, 1, 1, 8, "📈 30-Year Buy & Hold Projection")
    ws.row_dimensions[1].height = 40

    ln = 3
    ws.cell(row=ln, column=1, value="Assumptions:").font = section_font
    ln += 1
    assump = [
        ("Appreciation Rate", "3%", "0.00%"),
        ("Rent Growth Rate", "2%", "0.00%"),
        ("Expense Growth Rate", "2%", "0.00%"),
    ]
    for label, default, fmt in assump:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, default)
        ln += 1

    ln += 1
    headers = ["Year", "Property Value", "Monthly Rent", "Annual NOI",
               "Loan Balance", "Equity", "Cash Flow", "Cumulative CF"]
    set_header_row(ws, ln, headers, widths=col_widths)

    # Generate 30-year projection
    for yr in range(1, 31):
        r = ln + yr
        apply_cell(ws, r, 1, yr, font=value_font)
        ws.cell(row=r, column=1).number_format = "0"
        # Property value = purchase * (1+app)^yr
        apply_cell(ws, r, 2,
                   f'=ROUND("Deal Input"!C7*(1+$C$3)^{yr},2)',
                   font=value_font, number_format="$#,##0")
        # Monthly rent = base rent * (1+rent_growth)^yr
        apply_cell(ws, r, 3,
                   f'=ROUND("Deal Input"!C20*(1+$C$4)^{yr},2)',
                   font=value_font, number_format="$#,##0")
        # Annual NOI = rent*12 - expenses (simplified)
        apply_cell(ws, r, 4,
                   f'=ROUND(C{r}*12*(1-"Deal Input"!C21-"Deal Input"!C22)-(ROW(C{r})-9)*(ROW(C{r})-10)/0/0'
                   f'+("Deal Input"!C23+"Deal Input"!C24)*(1+$C$5)^{yr},2)',
                   font=value_font, number_format="$#,##0")
        # Loan balance = amortizing (simplified — remaining principal)
        apply_cell(ws, r, 5,
                   f'=ROUND(("Deal Input"!C7-"Deal Input"!C13*"Deal Input"!C7)*(1-(1-(1+"Deal Input"!C14/12)^(12*yr-12))/(1-(1+"Deal Input"!C14/12)^(12*"Deal Input"!C15)))*"Deal Input"!C15,2)',
                   font=value_font, number_format="$#,##0")
        # Equity = value - loan balance
        apply_cell(ws, r, 6,
                   f'=ROUND(B{r}-E{r},2)',
                   font=value_font, number_format="$#,##0")
        # Annual cash flow
        apply_cell(ws, r, 7,
                   f'=ROUND(D{r}-("Deal Input"!C7-"Deal Input"!C13*"Deal Input"!C7)*'
                   f'("Deal Input"!C14/12*(1+"Deal Input"!C14/12)^(12*"Deal Input"!C15))/((1+"Deal Input"!C14/12)^(12*"Deal Input"!C15)-1)*12,2)',
                   font=value_font, number_format="$#,##0")
        # Cumulative cash flow
        if yr == 1:
            apply_cell(ws, r, 8,
                       f'=G{r}',
                       font=value_font, number_format="$#,##0")
        else:
            apply_cell(ws, r, 8,
                       f'=G{r}+H{r-1}',
                       font=value_font, number_format="$#,##0")

    # Color header row
    r = ln + 31
    ws.cell(row=r, column=1, value="Summary:").font = section_font
    avg_items = [
        ("Avg Annual Cash Flow", f'=AVERAGE(H{ln+1}:H{ln+30})', "$#,##0"),
        ("Total Equity (Year 30)", f'=F{ln+30}', "$#,##0"),
        ("Total Cumulative Cash Flow", f'=H{ln+30}', "$#,##0"),
    ]
    for label, formula, fmt in avg_items:
        r += 1
        set_label(ws, r, 1, label)
        set_value(ws, r, 2, formula, fmt=fmt)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Deal Score / Summary
# ═══════════════════════════════════════════════════════════════════════════════
def build_deal_score(wb):
    ws = wb.create_sheet("Deal Score")
    ws.sheet_properties.tabColor = "8E44AD"

    col_widths = [40, 22, 22, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    merge_title(ws, 1, 1, 4, "📊 Deal Scorecard — One-Page Summary")
    ws.row_dimensions[1].height = 40

    ln = 3
    merge_section(ws, ln, 1, 4, "DEAL OVERVIEW")
    ln += 2
    overview = [
        ("Purchase Price", '="Deal Input"!C7', "$#,##0"),
        ("ARV", '="Deal Input"!C8', "$#,##0"),
        ("Rehab Budget", '="Deal Input"!C9', "$#,##0"),
        ("Total Cash Needed", '="BRRRR Analyzer"!C9', "$#,##0"),
        ("Monthly Rent", '="Deal Input"!C20', "$#,##0"),
        ("Cash Pulled Out (Refi)", '="BRRRR Analyzer"!B20', "$#,##0"),
    ]
    for label, formula, fmt in overview:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, formula, fmt=fmt)
        ln += 1

    ln += 1
    merge_section(ws, ln, 1, 4, "RETURN METRICS")
    ln += 2
    metrics = [
        ("Cash on Cash Return", '="BRRRR Analyzer"!B30', "0.00%"),
        ("BRRRR Quality", '="BRRRR Analyzer"!B35', None),
        ("Flip Profit", '="Flip Analyzer"!C20', "$#,##0"),
        ("Flip ROI", '="Flip Analyzer"!C21', "0.00%"),
        ("30yr Avg Annual CF", '="Buy & Hold 30yr"!C41', "$#,##0"),
    ]
    for label, formula, fmt in metrics:
        set_label(ws, ln, 1, label)
        set_value(ws, ln, 2, formula, fmt=fmt if fmt else None)
        ln += 1

    ln += 1
    merge_section(ws, ln, 1, 4, "DEAL QUALITY SCORE")
    ln += 2

    # Score display
    ws.merge_cells(start_row=ln, start_column=1, end_row=ln, end_column=2)
    cell = ws.cell(row=ln, column=1, value="")
    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # Score logic
    set_label(ws, ln+1, 1, "Score Conditions")
    set_value(ws, ln+1, 2, '=IF(AND("BRRRR Analyzer"!B20>0,"BRRRR Analyzer"!B30>=0.1),"GREEN 🟢",IF(OR("BRRRR Analyzer"!B20>0,"BRRRR Analyzer"!B30>=0.05),"YELLOW ⚠️","RED 🔴"))')

    # Score interpretation
    ln += 3
    ws.merge_cells(start_row=ln, start_column=1, end_row=ln+2, end_column=4)
    cell = ws.cell(row=ln, column=1,
                   value="🟢 GREEN: Strong deal — positive cash-out, CoC ≥ 10%\n"
                         "⚠️ YELLOW: Borderline — review assumptions closely\n"
                         "🔴 RED: Weak deal — cash-flow negative or poor returns")
    cell.font = Font(name="Calibri", color="666666", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[ln].height = 50

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: Refinance Scenario Builder
# ═══════════════════════════════════════════════════════════════════════════════
def build_refi_scenario(wb):
    ws = wb.create_sheet("Refi Scenarios")
    ws.sheet_properties.tabColor = "1ABC9C"

    col_widths = [32, 18, 18, 18, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    merge_title(ws, 1, 1, 6, "🔄 Refinance Scenario Builder — Compare LTV & Rates")
    ws.row_dimensions[1].height = 40

    ln = 3
    merge_section(ws, ln, 1, 6, "VARY REFINANCE LTV (Fixed ARV & Rate)")
    ln += 2

    headers = ["LTV %", "Max Loan", "Cash Out", "Cash Left In", "Monthly P&I", "CoC Return"]
    set_header_row(ws, ln, headers, widths=col_widths)

    for i in range(5):
        r = ln + 1 + i
        ltv_val = 0.65 + i * 0.025
        apply_cell(ws, r, 1, ltv_val, font=value_font, number_format="0.0%")
        # Max loan = ARV * LTV
        apply_cell(ws, r, 2,
                   f'=ROUND("Deal Input"!C8*${chr(65)}$' + str(r) + f',2)',
                   font=value_font, number_format="$#,##0")
        # Cash out = max loan - purchase*1.06 - rehab - holding
        apply_cell(ws, r, 3,
                   f'=ROUND(B{r}-"Deal Input"!C7*1.06-"Deal Input"!C9-"BRRRR Analyzer"!C7,2)',
                   font=value_font, number_format="$#,##0")
        # Cash left in
        apply_cell(ws, r, 4,
                   f'=ROUND("BRRRR Analyzer"!C9-C{r},2)',
                   font=value_font, number_format="$#,##0")
        # Monthly P&I on refi loan
        apply_cell(ws, r, 5,
                   f'=ROUND(PMT("Deal Input"!C17/12,"Deal Input"!C15*12,-B{r}),2)',
                   font=value_font, number_format="$#,##0")
        # CoC return after refi
        apply_cell(ws, r, 6,
                   f'=IF(D{r}>0,("BRRRR Analyzer"!B27-"BRRRR Analyzer"!B28)/D{r},"N/A")',
                   font=value_font, number_format="0.00%")

    ln += 8
    merge_section(ws, ln, 1, 6, "VARY ARV (Fixed LTV & Rate)")
    ln += 2
    set_header_row(ws, ln, headers, widths=col_widths)

    for i in range(5):
        r = ln + 1 + i
        arv_mult = 1.0 + (i - 2) * 0.1  # 0.8, 0.9, 1.0, 1.1, 1.2
        apply_cell(ws, r, 1, f'=ROUND("Deal Input"!C8*{arv_mult},0)',
                   font=value_font, number_format="$#,##0")
        apply_cell(ws, r, 2,
                   f'=ROUND(A{r}*"Deal Input"!C16,2)',
                   font=value_font, number_format="$#,##0")
        apply_cell(ws, r, 3,
                   f'=ROUND(B{r}-"Deal Input"!C7*1.06-"Deal Input"!C9-"BRRRR Analyzer"!C7,2)',
                   font=value_font, number_format="$#,##0")
        apply_cell(ws, r, 4,
                   f'=ROUND("BRRRR Analyzer"!C9-C{r},2)',
                   font=value_font, number_format="$#,##0")
        apply_cell(ws, r, 5,
                   f'=ROUND(PMT("Deal Input"!C17/12,"Deal Input"!C15*12,-B{r}),2)',
                   font=value_font, number_format="$#,##0")
        apply_cell(ws, r, 6,
                   f'=IF(D{r}>0,("BRRRR Analyzer"!B27-"BRRRR Analyzer"!B28)/D{r},"N/A")',
                   font=value_font, number_format="0.00%")

    # Highlight best scenario
    ln += 8
    ws.merge_cells(start_row=ln, start_column=1, end_row=ln, end_column=6)
    cell = ws.cell(row=ln, column=1,
                   value="💡 Tip: Look for the scenario with highest Cash Out AND positive Cash Left In for the best BRRRR result.")
    cell.font = note_font
    cell.alignment = Alignment(horizontal="left", vertical="center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7: Deal Criteria Rubric (Bonus)
# ═══════════════════════════════════════════════════════════════════════════════
def build_deal_criteria(wb):
    ws = wb.create_sheet("Deal Criteria")
    ws.sheet_properties.tabColor = "9B59B6"

    col_widths = [36, 22, 22, 22, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    merge_title(ws, 1, 1, 5, "🎯 Deal Criteria Rubric — Define Your Buy Box")
    ws.row_dimensions[1].height = 40

    ln = 3
    merge_section(ws, ln, 1, 5, "YOUR BUY BOX (Edit these thresholds to match your criteria)")
    ln += 2

    headers = ["Criteria", "Your Minimum", "Your Maximum", "Current Deal", "Pass/Fail"]
    set_header_row(ws, ln, headers, widths=col_widths)

    criteria = [
        ("Purchase Price", 0, 300000, '="Deal Input"!C7'),
        ("ARV", 0, 400000, '="Deal Input"!C8'),
        ("Rehab Budget", 0, 50000, '="Deal Input"!C9'),
        ("Cash on Cash Return", 0.08, 1.0, '="BRRRR Analyzer"!B30'),
        ("Cash Pulled Out", 0, 1000000, '="BRRRR Analyzer"!B20'),
        ("Max LTV", 0, 0.75, '="Deal Input"!C16'),
        ("Min Monthly Cash Flow", 100, 10000, '=("BRRRR Analyzer"!B27-"BRRRR Analyzer"!B28)/12'),
        ("Max Purchase Price", 0, 500000, '="Deal Input"!C7'),
        ("Min Rent/Price Ratio", 0.008, 1.0, '="Deal Input"!C20*12/"Deal Input"!C8'),
    ]

    for i, (crit, mn, mx, formula) in enumerate(criteria):
        r = ln + 1 + i
        apply_cell(ws, r, 1, crit, font=label_font, fill=white_fill,
                   alignment=Alignment(horizontal="left", vertical="center"))
        # Min
        apply_cell(ws, r, 2, mn, font=value_font)
        # Max
        apply_cell(ws, r, 3, mx, font=value_font)
        # Current deal value
        apply_cell(ws, r, 4, formula, font=value_font)
        # Pass/Fail formula
        apply_cell(ws, r, 5,
                   f'=IF(AND(D{r}>=B{r},D{r}<=C{r}),"✅ PASS","❌ FAIL")',
                   font=value_font)

    ln += len(criteria) + 3
    merge_section(ws, ln, 1, 5, "OVERALL BUY BOX RESULT")
    ln += 2
    set_label(ws, ln, 1, "All Criteria Met?")
    set_value(ws, ln, 2,
              f'=IF(COUNTIF(E{ln-len(criteria)-1}:E{ln-2},"❌ FAIL")=0,"✅ ALL PASS — BUY IT!","❌ Some criteria failed — review")')

    ln += 2
    cell = ws.cell(row=ln, column=1,
                   value="How to use: Set your minimum/maximum thresholds in columns B & C. The sheet auto-flags whether your deal meets your buy box.")
    cell.font = note_font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=ln, start_column=1, end_row=ln, end_column=5)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN BUILD
# ═══════════════════════════════════════════════════════════════════════════════
def build_spreadsheet(output_path):
    wb = openpyxl.Workbook()

    build_input_sheet(wb)
    build_brrrr_analyzer(wb)
    build_flip_analyzer(wb)
    build_buy_and_hold(wb)
    build_deal_score(wb)
    build_refi_scenario(wb)
    build_deal_criteria(wb)

    # Apply conditional formatting to BRRRR Analyzer score cells
    ws_brrrr = wb["BRRRR Analyzer"]
    green_font = Font(color="155724", bold=True)
    red_font = Font(color="721C24", bold=True)
    green_cell_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    red_cell_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    # Score row conditional formatting
    # Cash pulled out (row 20)
    for r in [20, 30, 35]:
        ws_brrrr.conditional_formatting.add(
            f'C{r}',
            CellIsRule(operator='greaterThan', formula=['0'],
                       font=green_font, fill=green_cell_fill)
        )
        ws_brrrr.conditional_formatting.add(
            f'C{r}',
            CellIsRule(operator='lessThanOrEqual', formula=['0'],
                       font=red_font, fill=red_cell_fill)
        )

    # Apply green/yellow/red on relevant score cells in Deal Score sheet
    ws_score = wb["Deal Score"]

    wb.save(output_path)
    print(f"✅ Spreadsheet saved to: {output_path}")
    return wb


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "BRRRR_Deal_Analyzer.xlsx"
    build_spreadsheet(out)
